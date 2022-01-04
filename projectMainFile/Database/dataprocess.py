import os
import sqlite3
import openpyxl
import subprocess

def sqlprompt(row, type):
    result = '''
    insert into course (
ser_no , /*FlowWater*/
co_chg ,
dpt_code ,
dptname,
cou_code,
class , /*班次*/
year , /*年級*/
credit ,
forh , /*半年  or全年 /
sel_code, /*必帶 , 必修 選修 */
cou_cname ,
cou_ename,
tea_cname,
clsrom_1 ,
clsrom_2 ,
daytime, /*時間，str模式*/
mark, /*備註*/
co_tp, /*學校excel 2 == 共同必修 */
co_gmark, /*用途不明*/
clsrom_3 ,
clsrom_4 ,
clsrom_5 ,
clsrom_6 ,
co_select , /*加選*/
tlec, /*正課時數*/
tlab, /*實驗時數*/
time0 ,
time1 ,
time2 ,
time3 ,
time4 ,
time5 ,
time6 ,
time7 ,
time8 ,
time9 ,
time10 ,
time11 ,
time12 ,
time13 ,
time14 ,
time15 ,
time16 ,
time17 ,
time18 ,
time19 , /* need to parse it from excel*/
gradelimit, /* 0==不限, 1==限大1 2以上==限大n以上*/
majorlimit,/*限本系*/
category , /* 根據放在哪個excel檔案決定 0==共同 1==通識 2==本係*/
category2 , /*擴充用*/
noOpen, /*初選不開放*/
libcat1, /*通識分類*/
libcat2
)
values ('''
    row = [str(row[i]) for i in range(0,26)]
    for i in range(0, 26):
        row[i] = row[i].replace(" ","")
        row[i] = row[i].replace("_x0000_", "")
        row[i] = row[i].replace("\"", "")

    clsromlst = [14,15,19,20,21,22]
    clasromstr = [""]*6
    for index in clsromlst:
        if (len(row[index]) > 3):
            clasromstr.append(row[index])
    for i in range(0,13):
        try:
            temp = int(row[i])
            result += str(temp) + ","
        except:
            result += "\"" + row[i] + "\"" + ","
    for i in range(0, 2):
        result +=  "\"" + clasromstr[i] + "\"" + ","
    for i in range(15, 18):
        try:
            temp = int(row[i])
            result += str(temp) + ","
        except:
            result += "\"" + row[i] + "\"" + ","
    for i in range(2, 6):
        result +=  "\"" + clasromstr[i] + "\"" + ","
    for i in range(22, 26):
        try:
            temp = int(row[i])
            result += str(temp) + ","
        except:
            result += "\"" + row[i] + "\"" + ","
    timeList = getTimeList(row[15])
    for i in range(0,20):
        result += str(timeList[i]) + ","
    if "限大一" in row[16]:
        result += "1,"
        if "限大二" in row[16]:
            result += "2,"
            if "限大三" in row[16]:
                    result += "3,"
                    if "限大四" in row[16]:
                        result += "4,"
                    else:
                        result+= "0,"
    if "限本系" in row[16]:
        result += "true,"
    else:
        result += "false,"
    result += type[0] + ","
    if type[0] == '2' and "通識" in row[16]:
        result += '1' + ","
    else:
        result += '-1' + ","
    if "初選不開放" in row[16]:
        result += "true,"
    else:
        result += "false,"
    libcat = []
    if "文學與藝術" in row[16]:
        libcat.append(1)
    if "歷史思維" in row[16]:
        libcat.append(2)
    if "世界文明" in row[16]:
        libcat.append(3)
    if "哲學與道德" in row[16]:
        libcat.append(4)
    if "公民意識" in row[16]:
        libcat.append(5)
    if "量化分析" in row[16]:
        libcat.append(6)
    if "物質科學" in row[16]:
        libcat.append(7)
    if "生命科學" in row[16]:
        libcat.append(8)
    libcat.append(0)
    libcat.append(0)
    result += f"{libcat[0]}, {libcat[1]}"
    result += ");"
    return result

def getTimeList(st):
    dir = os.getcwd()+"\\a"
    return [str(int(i)) for i in subprocess.Popen(dir + " " + st, stdout=subprocess.PIPE).communicate()[0].split()]

dir = os.getcwd()
db = sqlite3.connect(dir + '\\course.db');
cursor = db.cursor()
workbook = openpyxl.load_workbook(dir + "\\CourseExcel\\" + input("input the name of the excel file"))
sheet = workbook.worksheets[0]
type =  [input("Category: 0 共同 1 通識 2 系上課"), input("Category2: 0 共同 1 通識 2 系上課")]

#https://ithelp.ithome.com.tw/articles/10246377
for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
    print(sqlprompt(row, type))
    cursor.execute(sqlprompt(row, type))
db.commit()
